import pandas as pd
from pathlib import Path
from typing import List
from datetime import datetime
from app.models import RequestInDB
from app.config import settings
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
import shutil

logger = logging.getLogger(__name__)


def export_requests_to_excel(requests: List[RequestInDB], loai_mau: int) -> Path:
    """
    Export requests to Excel file matching the form template

    Args:
        requests: List of requests to export
        loai_mau: Form template number (6-10 for white/yellow plates)

    Returns:
        Path to exported Excel file
    """
    # Create output directory
    output_dir = settings.REQUESTS_DIR / "exports"
    output_dir.mkdir(exist_ok=True)

    # Copy template file
    template_file = Path(__file__).parent.parent.parent / "update_2911" / f"Mẫu {loai_mau}.xlsx"

    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_file}")

    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"Mau_{loai_mau}_Export_{timestamp}.xlsx"

    # Copy template
    shutil.copy(template_file, output_file)

    # Load workbook
    wb = load_workbook(output_file)
    ws = wb.active

    # Find data start row (after headers)
    data_start_row = 4  # Row 4 is headers, Row 5 starts data

    # Unmerge all cells in data area to avoid MergedCell issues
    # This allows us to write data to any cell without conflicts
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        # Only unmerge cells in data rows (row 5 onwards)
        if merged_range.min_row >= data_start_row + 1:
            ws.unmerge_cells(str(merged_range))
            logger.debug(f"Unmerged range: {merged_range}")

    # Prepare and write data
    data_rows = _prepare_export_data(requests, loai_mau)

    for idx, row_data in enumerate(data_rows, start=data_start_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx)

            try:
                cell.value = value
                # Apply basic formatting
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            except Exception as e:
                logger.warning(f"Error writing to cell at row {idx}, col {col_idx}: {e}")
                continue

    # Save workbook
    wb.save(output_file)
    logger.info(f"Exported {len(requests)} requests to {output_file}")
    return output_file


def _prepare_export_data(requests: List[RequestInDB], loai_mau: int) -> List[List]:
    """Prepare data rows for export based on form type - matching EXACT Excel template structure"""
    data_rows = []

    for idx, req in enumerate(requests, 1):
        if loai_mau == 6:
            # Mẫu 6: Xe và chủ xe đúng
            # Cols: STT, Biển số, Màu biển, Loại xe, Chủ xe, Địa chỉ, Số khung, Số máy,
            #       SĐT chủ xe, CCCD chủ xe, Ngày cấp CCCD, GPLX, Ngày cấp GPLX, Cơ quan cấp, Tình trạng, Ghi chú
            row = [
                idx,  # Col 1: STT
                req.bien_so or "",  # Col 2: Biển số
                req.mau_bien or "",  # Col 3: Màu biển
                req.loai_xe or "",  # Col 4: Loại xe
                req.chu_xe or "",  # Col 5: Chủ xe
                req.dia_chi_chu_xe or "",  # Col 6: Địa chỉ thường trú, địa chỉ hiện tại
                req.so_khung or "",  # Col 7: Số khung
                req.so_may or "",  # Col 8: Số máy
                req.so_dien_thoai_chu_xe or "",  # Col 9: Số Điện thoại
                req.ma_so_thue_chu_xe or "",  # Col 10: Số CCCD/mã số thuế
                req.ngay_cap_cccd_chu_xe or "",  # Col 11: Ngày cấp CCCD
                req.so_gplx_chu_xe or "",  # Col 12: Số GPLX
                req.ngay_cap_gplx_chu_xe or "",  # Col 13: Ngày cấp GPLX
                req.co_quan_cap_gplx_chu_xe or "",  # Col 14: Cơ quan cấp GPLX
                req.tinh_trang_xe or "",  # Col 15: Tình trạng phương tiện
                req.ghi_chu or ""  # Col 16: Ghi chú
            ]

        elif loai_mau == 7:
            # Mẫu 7: Có chủ xe, không có xe
            # Cols: STT, Biển số, Màu biển, Loại xe, Chủ xe, Địa chỉ, SĐT chủ xe, CCCD chủ xe, Ngày cấp CCCD,
            #       GPLX, Ngày cấp GPLX, Cơ quan cấp, Tên người mua VÀ ĐỊA CHỈ, Bản sao, SĐT người mua, CCCD người mua, Ngày cấp CCCD người mua
            row = [
                idx,  # Col 1: STT
                req.bien_so or "",  # Col 2: Biển số
                req.mau_bien or "",  # Col 3: Màu biển
                req.loai_xe or "",  # Col 4: Loại xe
                req.chu_xe or "",  # Col 5: Chủ xe
                req.dia_chi_chu_xe or "",  # Col 6: Địa chỉ thường trú, địa chỉ hiện tại
                req.so_dien_thoai_chu_xe or "",  # Col 7: Số Điện thoại chủ xe
                req.ma_so_thue_chu_xe or "",  # Col 8: Số CCCD/mã số thuế
                req.ngay_cap_cccd_chu_xe or "",  # Col 9: Ngày cấp CCCD
                req.so_gplx_chu_xe or "",  # Col 10: Số GPLX
                req.ngay_cap_gplx_chu_xe or "",  # Col 11: Ngày cấp GPLX
                req.co_quan_cap_gplx_chu_xe or "",  # Col 12: Cơ quan cấp GPLX
                # Col 13: Tên người mua VÀ ĐỊA CHỈ (combined)
                f"{req.ten_nguoi_mua or ''} - {req.dia_chi_nguoi_mua or ''}" if req.ten_nguoi_mua else "",
                req.ban_sao_chuyen_nhuong or "",  # Col 14: Bản sao chứng từ
                req.so_dien_thoai_nguoi_mua or "",  # Col 15: Số Điện thoại người mua
                req.so_cccd_nguoi_mua or "",  # Col 16: Số CCCD người mua
                req.ngay_cap_cccd_nguoi_mua or ""  # Col 17: Ngày cấp CCCD người mua
            ]

        elif loai_mau == 8:
            # Mẫu 8: Có xe, không có chủ xe
            # Cols: STT, Biển số, Màu biển, Loại xe, Chủ xe, Địa chỉ, SĐT chủ xe, CCCD chủ xe, Ngày cấp CCCD,
            #       GPLX người đang sử dụng, Ngày cấp GPLX chủ xe, Cơ quan cấp, Tên người mua VÀ ĐỊA CHỈ,
            #       SĐT người mua, CCCD người mua, Ngày cấp CCCD người mua, Bản sao
            row = [
                idx,  # Col 1: STT
                req.bien_so or "",  # Col 2: Biển số
                req.mau_bien or "",  # Col 3: Màu biển
                req.loai_xe or "",  # Col 4: Loại xe
                req.chu_xe or "",  # Col 5: Chủ xe
                req.dia_chi_chu_xe or "",  # Col 6: Địa chỉ thường trú, địa chỉ hiện tại
                req.so_dien_thoai_chu_xe or "",  # Col 7: Số Điện thoại chủ xe
                req.ma_so_thue_chu_xe or "",  # Col 8: Số CCCD/mã số thuế
                req.ngay_cap_cccd_chu_xe or "",  # Col 9: Ngày cấp CCCD
                req.so_gplx_chu_xe or "",  # Col 10: Số GPLX của người đang sử dụng xe
                req.ngay_cap_gplx_chu_xe or "",  # Col 11: Ngày cấp GPLX của chủ xe
                req.co_quan_cap_gplx_chu_xe or "",  # Col 12: Cơ quan cấp GPLX
                # Col 13: Tên người mua VÀ ĐỊA CHỈ (combined)
                f"{req.ten_nguoi_mua or ''} - {req.dia_chi_nguoi_mua or ''}" if req.ten_nguoi_mua else "",
                req.so_dien_thoai_nguoi_mua or "",  # Col 14: Số Điện thoại người mua
                req.so_cccd_nguoi_mua or "",  # Col 15: Số CCCD/mã số thuế người mua
                req.ngay_cap_cccd_nguoi_mua or "",  # Col 16: Ngày cấp CCCD người mua
                req.ban_sao_chuyen_nhuong or ""  # Col 17: Bản sao chứng từ
            ]

        elif loai_mau == 9:
            # Mẫu 9: Không có xe, không có chủ xe
            # Cols: STT, Biển số, Màu biển, Loại xe, Chủ xe, Địa chỉ, SĐT chủ xe, CCCD chủ xe, Ngày cấp CCCD,
            #       Tên người mua VÀ ĐỊA CHỈ, SĐT người mua, CCCD người mua, Ngày cấp CCCD người mua, Bản sao
            row = [
                idx,  # Col 1: STT
                req.bien_so or "",  # Col 2: Biển số
                req.mau_bien or "",  # Col 3: Màu biển
                req.loai_xe or "",  # Col 4: Loại xe
                req.chu_xe or "",  # Col 5: Chủ xe
                req.dia_chi_chu_xe or "",  # Col 6: Địa chỉ thường trú, địa chỉ hiện tại
                req.so_dien_thoai_chu_xe or "",  # Col 7: Số Điện thoại của chủ xe
                req.ma_so_thue_chu_xe or "",  # Col 8: Số CCCD/mã số thuế của chủ xe
                req.ngay_cap_cccd_chu_xe or "",  # Col 9: Ngày cấp CCCD của chủ xe
                # Col 10: Tên người mua VÀ ĐỊA CHỈ (combined)
                f"{req.ten_nguoi_mua or ''} - {req.dia_chi_nguoi_mua or ''}" if req.ten_nguoi_mua else "",
                req.so_dien_thoai_nguoi_mua or "",  # Col 11: Số Điện thoại người mua
                req.so_cccd_nguoi_mua or "",  # Col 12: Số CCCD/mã số thuế người mua
                req.ngay_cap_cccd_nguoi_mua or "",  # Col 13: Ngày cấp CCCD người mua
                req.ban_sao_chuyen_nhuong or ""  # Col 14: Bản sao chứng từ
            ]

        elif loai_mau == 10:
            # Mẫu 10: Xe không nằm trong danh sách
            # Cols: STT, Biển số, Màu biển, Loại xe, Chủ xe VÀ ĐỊA CHỈ, Người đang sử dụng VÀ ĐỊA CHỈ,
            #       SĐT chủ xe, CCCD chủ xe, Ngày cấp CCCD chủ xe, GPLX, Ngày cấp GPLX, Cơ quan cấp,
            #       Người bán VÀ ĐỊA CHỈ, SĐT người bán VÀ người đang sử dụng, CCCD người bán VÀ người đang sử dụng,
            #       Ngày cấp CCCD người bán VÀ người đang sử dụng, Bản sao
            row = [
                idx,  # Col 1: STT
                req.bien_so or "",  # Col 2: Biển số
                req.mau_bien or "",  # Col 3: Màu biển
                req.loai_xe or "",  # Col 4: Loại xe
                # Col 5: Chủ xe VÀ ĐỊA CHỈ (combined)
                f"{req.chu_xe or ''} - {req.dia_chi_chu_xe or ''}" if req.chu_xe else "",
                # Col 6: Người đang sử dụng VÀ ĐỊA CHỈ (combined)
                f"{req.ten_nguoi_dang_su_dung or ''} - {req.dia_chi_nguoi_dang_su_dung or ''}" if req.ten_nguoi_dang_su_dung else "",
                req.so_dien_thoai_chu_xe or "",  # Col 7: Số Điện thoại của chủ xe
                req.ma_so_thue_chu_xe or "",  # Col 8: Số CCCD/mã số thuế của chủ xe
                req.ngay_cap_cccd_chu_xe or "",  # Col 9: Ngày cấp CCCD/mã số thuế của chủ xe
                req.so_gplx_chu_xe or "",  # Col 10: Số GPLX của chủ xe
                req.ngay_cap_gplx_chu_xe or "",  # Col 11: Ngày cấp GPLX của chủ xe
                req.co_quan_cap_gplx_chu_xe or "",  # Col 12: Cơ quan cấp GPLX của chủ xe
                # Col 13: Tên người bán xe VÀ ĐỊA CHỈ (combined)
                f"{req.ten_nguoi_ban or ''} - {req.dia_chi_nguoi_ban or ''}" if req.ten_nguoi_ban else "",
                # Col 14: SĐT của người bán xe VÀ SĐT người đang sử dụng (combined)
                f"{req.so_dien_thoai_nguoi_ban or ''} - {req.so_dien_thoai_nguoi_mua or ''}" if (req.so_dien_thoai_nguoi_ban or req.so_dien_thoai_nguoi_mua) else "",
                # Col 15: CCCD người bán xe VÀ người đang sử dụng (combined)
                f"{req.so_cccd_nguoi_ban or ''} - {req.so_cccd_nguoi_mua or ''}" if (req.so_cccd_nguoi_ban or req.so_cccd_nguoi_mua) else "",
                # Col 16: Ngày cấp CCCD người bán xe VÀ người đang sử dụng (combined)
                f"{req.ngay_cap_cccd_nguoi_ban or ''} - {req.ngay_cap_cccd_nguoi_mua or ''}" if (req.ngay_cap_cccd_nguoi_ban or req.ngay_cap_cccd_nguoi_mua) else "",
                req.ban_sao_chuyen_nhuong or ""  # Col 17: Bản sao chứng từ
            ]

        else:
            # Default fallback (for mẫu 1-5)
            row = [
                idx,
                req.bien_so or "",
                req.loai_xe or "",
                req.chu_xe or "",
                req.dia_chi_chu_xe or "",
                req.so_dien_thoai_chu_xe or "",
                req.ghi_chu or ""
            ]

        data_rows.append(row)

    return data_rows


def _format_excel_header(worksheet, loai_mau: int):
    """Format Excel header to match template - Not needed if using template copy"""
    pass
